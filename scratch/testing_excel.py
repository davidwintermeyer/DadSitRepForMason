import openpyxl
from openpyxl import load_workbook

# https://stackoverflow.com/questions/33541692/how-to-find-the-last-row-in-a-column-using-openpyxl-normal-workbook
from constants import file_constants

SIT_REP_FILE_PATH = 'Covid-19_SitRep_Data-01-19-2021.xlsx'

def get_last_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    return ws.max_row

def get_cell(column_letter: str, row_number: int) -> str:
    return column_letter + str(row_number)

def update_sitrep():

    # Load the workbook
    wb = load_workbook(filename=SIT_REP_FILE_PATH)

    # Pull the sheet
    sheet = wb[file_constants.SHEET_NAME]

    # Add a row at the end
    last_row = get_last_row(sheet) + 1
    sheet.insert_rows(last_row)

    # Go through each column
    sheet[get_cell('A', last_row)] = 'Testing'
    print(sheet[get_cell('A', last_row)].value)

    # Save the workbook
    wb.save(filename=SIT_REP_FILE_PATH)

update_sitrep()


