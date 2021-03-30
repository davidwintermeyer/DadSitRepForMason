import datetime
import json

## email_template_generator_handler invoked with event:
import traceback

from openpyxl import load_workbook

from constants.email_constants import EMAIL_RECIPIENTS
from constants.file_constants import get_s3_key, PRIVATE_S3_BUCKET_NAME, get_file_name
from email_template_generator.email_template_processor import get_email_text_paragraphs_in_list, get_email_html
from util.date_util import get_days_ago
from util.lambda_util import get_report_date_time
from util.s3_util import download_file, upload_workbook
from util.ses_util import send_report_as_attachment


def get_bucket_and_key_from_event(event: dict) -> []:
    individual_record = event["Records"][0]
    s3_record = individual_record["s3"]
    bucket = s3_record["bucket"]["name"]
    key = s3_record["object"]["key"]
    return [bucket, key]


def get_report_date_from_s3_event(event):
    individual_record_time_str_unformatted = event["Records"][0]["eventTime"]
    # [ERROR] ValueError: time data '2021-03-17T04:55:56.283Z' does not match format '%Y-%m-%dT%H:%M:%SZ'
    # we want to delete those milliseconds
    split_string = individual_record_time_str_unformatted.split(".", 1)
    individual_record_time_str_formatted = split_string[0] + 'Z'
    report_date_time = get_report_date_time(individual_record_time_str_formatted)
    # If its before 5pm, treat is as previous day
    days_ago = 0
    if report_date_time.time() < datetime.time(17, 0):
        days_ago = 1
    return get_days_ago(report_date_time.date(), days_ago)


def get_workbook_file_only_given_s3_event(event):
    bucket_and_key = get_bucket_and_key_from_event(event)
    source_bucket = bucket_and_key[0]
    source_key = bucket_and_key[1]

    # Load the workbook
    print("downloading file from bucket: " + source_bucket + " with key: " + source_key)
    # https://stackoverflow.com/questions/17195569/using-a-variable-in-a-try-catch-finally-statement-without-declaring-it-outside
    try:
        return download_file(bucket=source_bucket, key=source_key)
    except Exception as exc:
        raise RuntimeError('Failed to file from bucket: ' + source_bucket + ' with key: ' + source_key) from exc


def get_text_as_string(text_paragraphs_in_list):
    text = ''
    for paragraph in text_paragraphs_in_list:
        text += paragraph
        text += '\n'
    return text


def get_text_as_html(text_paragraphs_in_list):
    # Wrap each paragraph in its own <p> </p>
    # First sentence
    text = ''
    for paragraph in text_paragraphs_in_list:
        text += '<p>'
        lines = paragraph.splitlines()
        is_first = True
        # for the one liners, don't make them bold.
        if len(lines) == 1:
            is_first = False
        for line in lines:
            if is_first:
                text += '<b>'
                text += line
                text += '</b>'
                is_first = False
            else:
                text += line
            text += '<br>'
            text += '</p>'
    return text


def get_subject(report_date):
    date_str = report_date.strftime('%m/%d/%Y')

    return "Mason Covid Sitrep Report Email Text: " + date_str

def email_template_generator_handler(event, context):
    print("email_template_generator_handler invoked with event: " + json.dumps(event))

    report_date = get_report_date_from_s3_event(event)
    new_report_s3_key = get_s3_key(report_date)

    file = get_workbook_file_only_given_s3_event(event)
    workbook_to_upload_to_s3 = load_workbook(file)

    try:
        # Write the report to the non-public bucket
        upload_workbook(workbook=workbook_to_upload_to_s3, bucket=PRIVATE_S3_BUCKET_NAME, key= new_report_s3_key)
        print("Report with data uploaded to website written to private bucket. Bucket: " + PRIVATE_S3_BUCKET_NAME + " key: " + new_report_s3_key)
    except Exception as exc:
        error_message = "Error copying sitrep to private bucket! Bucket: " + PRIVATE_S3_BUCKET_NAME + " key: " + new_report_s3_key
        raise RuntimeError(error_message) from exc


    data_only_workbook = load_workbook(file, data_only=True)
    sheet = data_only_workbook.active

    try:
        print("Processing email template for report with report_date: " + str(report_date))

        email_html = get_email_html(sheet, report_date)

        # The HTML body of the email.
        body_html = """\
        <html>
        <head></head>
        <body>
        <h1>Please see the following auto-generated email template</h1>
        """
        body_html += email_html
        body_html += """
        </body>
        </html>
        """

        report_local_path = '/tmp/' + get_file_name(report_date)
        print("saving file locally to path: " + report_local_path)
        workbook_to_upload_to_s3.save(filename=report_local_path)

        email_recipients = EMAIL_RECIPIENTS
        send_report_as_attachment(report_local_path=report_local_path, email_recipients=email_recipients, subject=get_subject(report_date), body_text=email_html, body_html=body_html)

    except Exception as exc:
        traceback.print_exc()
        print(exc)
        error_message = "Processing email template for report with report_date: " + str(report_date)
        raise RuntimeError(error_message, exc) from exc
