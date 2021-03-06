from datetime import date, time

from openpyxl import load_workbook, Workbook

from constants import file_constants, sitrep_column_constants
from constants.file_constants import COLUMN_NAME_TO_STYLE_NAME_DICT, STYLE_NAME_TO_STYLE_DICT, get_file_name
from constants.sitrep_column_constants import column_title_to_letter_dicts, NO_DATA_COLUMNS, NO_DATA
from dependencies.csse_github import get_csse_data
from util.excel_util import get_last_row, get_cell
# Given an input file path to an existing sitrep excel file,
# adds a new record for the report_date input, and writes the file to output_file_path
# TODO: Validations?
# If you want to overwrite the same file, set input_file_path and output_file_path to be the same.
from dependencies.vhd_soda import get_vdh_data

#
from util.s3_util import download_file, upload_workbook

# returns path of the file
def generate_report_for_day_s3(s3bucket: str, previous_report_s3_key: str, new_report_s3_key: str, report_date: date, report_time: time) -> str:

    # Load the workbook
    print("downloading file from bucket: " + s3bucket + " with key: " + previous_report_s3_key)
    # https://stackoverflow.com/questions/17195569/using-a-variable-in-a-try-catch-finally-statement-without-declaring-it-outside
    try:
        file = download_file(bucket=s3bucket, key=previous_report_s3_key)
    except Exception as exc:
        raise RuntimeError('Failed to get sit from previous day. Key: ' + previous_report_s3_key) from exc

    wb = load_workbook(file)

    # Should update the wb itself
    update_workbook_with_new_record(wb=wb, report_date=report_date, report_time=report_time)

    # Save the workbook
    local_filepath = '/tmp/' + get_file_name(report_date)
    print("saving file locally to path: " + local_filepath)
    wb.save(filename=local_filepath)

    print("uploading file to bucket: " + s3bucket + " with key: " + new_report_s3_key)
    upload_workbook(workbook=wb, bucket=s3bucket, key=new_report_s3_key)

    return local_filepath


def generate_report_for_day_write_file_locally(input_file_path: str, output_file_path: str, report_date: date, report_time: time):

    # Load the workbook
    wb = load_workbook(filename=input_file_path)

    # Should update the wb itself
    update_workbook_with_new_record(wb=wb, report_date=report_date, report_time=report_time)

    # Save the workbook
    wb.save(filename=output_file_path)

def update_workbook_with_new_record(wb: Workbook, report_date: date, report_time: time):

    column_title_to_value_dict = {}
    column_title_to_value_dict[sitrep_column_constants.DATE_COLUMN] = report_date
    column_title_to_value_dict[sitrep_column_constants.TIME_COLUMN] = report_time

    # csse data
    csse_data_dict = get_csse_data(report_date)
    column_title_to_value_dict.update(csse_data_dict)

    # vdh Data
    vdh_dict = get_vdh_data(report_date)
    column_title_to_value_dict.update(vdh_dict)

    # Write in No Data
    for column_title in NO_DATA_COLUMNS :
        column_title_to_value_dict[column_title] = NO_DATA

    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html
    # check if style names are registered, and if not, register
    workbook_style_names = wb.style_names
    for style_name in STYLE_NAME_TO_STYLE_DICT.keys():
        if style_name not in workbook_style_names:
            wb.add_named_style(STYLE_NAME_TO_STYLE_DICT[style_name])

    # Pull the sheet
    sheet = wb[file_constants.SHEET_NAME]

    # # Add a row at the end
    # last_row_number = get_last_row(sheet) + 1
    # sheet.insert_rows(last_row_number)
    last_row_number = file_constants.get_row_number(report_date)
    # With the formulas, now we don't look fro the last row number, we have a hardcoded map.

    # Update the last row with values for report_date
    for column_title, column_letter in column_title_to_letter_dicts.items():
        cell = sheet[get_cell(column_letter, last_row_number)]
        if column_title in column_title_to_value_dict:
            cell.value = column_title_to_value_dict[column_title]
        if column_title in COLUMN_NAME_TO_STYLE_NAME_DICT:
            cell.style = COLUMN_NAME_TO_STYLE_NAME_DICT[column_title]